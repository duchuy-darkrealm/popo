from openpyxl import Workbook
from openpyxl import load_workbook

columns = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]

class MyOpenpyxl:
    data = []
    
    def readValue(self,filename):
        ws = load_workbook(filename).active
        i = 0
        labels = []
        while ws[columns[i] + "1"].value != None:
            label = {}
            label["column"] = columns[i]
            label["value"] = ws[columns[i] + "1"].value
            labels.append(label)
            i+=1

        row = 2
        result = []
        while ws["A"+str(row)].value!=None:
            item = {}
            for label in labels:
                item[label["value"]] = ws[label["column"] + str(row)].value
                
            result.append(item)
            row += 1
        data = result
        return result
    
        
